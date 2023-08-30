#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from datetime import datetime

def submit_leave():
    employee_id = employee_id_entry.get()
    employee_name = employee_name_entry.get()
    institution = institution_var.get()
    leave_type = leave_type_var.get()
    date = date_entry.get()
    hours = hours_entry.get()
    reason = reason_text.get("1.0", "end-1c")

    if not employee_id or not employee_name or not institution or not leave_type or not date or not hours or not reason:
        messagebox.showerror("錯誤", "請填寫所有欄位")
        return

    try:
        hours = float(hours)
    except ValueError:
        messagebox.showerror("錯誤", "請輸入有效的時數")
        return

    filename = "leave_records.xlsx"
    try:
        wb = load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["員編", "姓名", "所屬機構", "假別", "日期", "時數", "事由"])

    ws.append([employee_id, employee_name, institution, leave_type, date, hours, reason])
    wb.save(filename)
    messagebox.showinfo("成功", "資料已儲存到Excel檔案")
    
    employee_id_entry.delete(0, "end")
    employee_name_entry.delete(0, "end")
    institution_var.set(institution_options[0])  # 預設選第一個選項
    leave_type_var.set(leave_type_options[0])    # 預設選第一個選項
    date_entry.delete(0, "end")
    hours_entry.delete(0, "end")
    reason_text.delete("1.0", "end")

app = tk.Tk()
app.title("請假紀錄")

# 員編
employee_id_frame = tk.Frame(app)
employee_id_frame.pack(fill="x", padx=10, pady=5)
employee_id_label = tk.Label(employee_id_frame, text="員編:")
employee_id_label.pack(side="left")
employee_id_entry = tk.Entry(employee_id_frame)
employee_id_entry.pack(side="left")

# 所屬機構
institution_frame = tk.Frame(app)
institution_frame.pack(fill="x", padx=10, pady=5)
institution_label = tk.Label(institution_frame, text="所屬機構:")
institution_label.pack(side="left")
institution_var = tk.StringVar()
institution_options = ["府城", "鴻康", "謙益", "寬澤"]
institution_menu = tk.OptionMenu(institution_frame, institution_var, *institution_options)
institution_menu.pack(side="left")

# 姓名
employee_name_frame = tk.Frame(app)
employee_name_frame.pack(fill="x", padx=10, pady=5)
employee_name_label = tk.Label(employee_name_frame, text="姓名:")
employee_name_label.pack(side="left")
employee_name_entry = tk.Entry(employee_name_frame)
employee_name_entry.pack(side="left")

# 假別
leave_type_frame = tk.Frame(app)
leave_type_frame.pack(fill="x", padx=10, pady=5)
leave_type_label = tk.Label(leave_type_frame, text="假別:")
leave_type_label.pack(side="left")
leave_type_var = tk.StringVar()
leave_type_options = ["事假", "病假", "喪假", "休假", "婚假", "生理假", "特休", "公假"]
leave_type_menu = tk.OptionMenu(leave_type_frame, leave_type_var, *leave_type_options)
leave_type_menu.pack(side="left")

# 日期
date_frame = tk.Frame(app)
date_frame.pack(fill="x", padx=10, pady=5)
date_label = tk.Label(date_frame, text="日期:")
date_label.pack(side="left")
date_entry = tk.Entry(date_frame)
date_entry.pack(side="left")

# 時數
hours_frame = tk.Frame(app)
hours_frame.pack(fill="x", padx=10, pady=5)
hours_label = tk.Label(hours_frame, text="時數:")
hours_label.pack(side="left")
hours_entry = tk.Entry(hours_frame)
hours_entry.pack(side="left")

# 事由
reason_frame = tk.Frame(app)
reason_frame.pack(fill="x", padx=10, pady=5)
reason_label = tk.Label(reason_frame, text="事由:")
reason_label.pack(side="left")
reason_text = tk.Text(reason_frame, height=5, width=30)
reason_text.pack(side="left")

# 送出按鈕
submit_button = tk.Button(app, text="送出", command=submit_leave)
submit_button.pack(padx=10, pady=10)

app.mainloop()


# In[ ]:




